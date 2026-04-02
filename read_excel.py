import openpyxl
from openpyxl.utils import get_column_letter
import json

# Read the Excel file
wb = openpyxl.load_workbook('c:/Users/basel/OneDrive/Desktop/Not_Messaged_List.xlsx')
ws = wb.active

# Get all rows
rows = []
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    rows.append(row)
    if idx >= 4:  # Just first 5 rows
        break

# Print headers
print("=== Excel Content ===")
print(f"Dimensions: {ws.dimensions}")
print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
print("\n=== First 5 Rows ===")
for i, row in enumerate(rows):
    print(f"Row {i+1}: {row}")
